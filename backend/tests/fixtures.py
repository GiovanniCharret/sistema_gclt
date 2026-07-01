"""Gerador de `.xlsx`-fixture da aba `Preenchimento` (openpyxl) — Bloco D.

Por que existe: os testes de parsing/validação (§13 da spec) precisam fabricar planilhas
com a aba `Preenchimento` (cabeçalho na linha 2, dados a partir da linha 3), parametrizáveis
por linha/célula para acionar cada regra — sem depender de um arquivo real (que não existe).
"""

# `io.BytesIO` devolve o .xlsx em memória (bytes), como chega no upload.
import io
# `openpyxl` monta a planilha.
import openpyxl

# Subconjunto realista de colunas (as de identificação/localização + 2 tipologias),
# suficiente para exercitar as regras. O parser mapeia por NOME, então a ordem/qtd é livre.
CABECALHO_PADRAO = [
    "Distribuidora", "Tipo de Atendimento", "Número ODI",
    "Número da Unidade Consumidora", "Código IBGE do Município", "Município",
    "UF", "Latitude", "Longitude", "Data de Energização da UC",
    "Tipo de Comunidade", "Enquadramento do beneficiário",
    "0 - Não é prioridade", "I - Baixa renda",
]


def gerar_xlsx(linhas, cabecalho=None, aba="Preenchimento"):
    """Cria um .xlsx (bytes) com a aba dada, cabeçalho na linha 2 e dados da linha 3.

    Entrada: `linhas` (lista de dicts {nome_coluna: valor}), `cabecalho` (lista de nomes;
             None usa o padrão) e `aba` (nome da planilha).
    Fase 1: cria o workbook e nomeia a aba.
    Fase 2: escreve o cabeçalho na linha 2 (linha 1 fica como cabeçalho de seção, em branco).
    Fase 3: escreve cada linha de dados a partir da linha 3 (célula vazia p/ chave ausente).
    Saída: os bytes do arquivo .xlsx.
    """
    # Cabeçalho efetivo.
    cabecalho = cabecalho if cabecalho is not None else CABECALHO_PADRAO
    # Fase 1: workbook + aba.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = aba
    # Fase 2: cabeçalho real na linha 2 (linha 1 reservada p/ cabeçalho de seção).
    for c, nome in enumerate(cabecalho, start=1):
        ws.cell(row=2, column=c, value=nome)
    # Fase 3: dados a partir da linha 3.
    for i, linha in enumerate(linhas):
        for c, nome in enumerate(cabecalho, start=1):
            ws.cell(row=3 + i, column=c, value=linha.get(nome))
    # Saída: serializa em bytes.
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_xlsx_dominios(colunas, aba="Dominios"):
    """Cria um .xlsx (bytes) com a aba `Dominios`: cabeçalho na linha 1, valores abaixo.

    Entrada: `colunas` (dict {nome_coluna: [valores]}) e `aba`.
    Fase 1: cria o workbook e nomeia a aba.
    Fase 2: escreve cada cabeçalho na linha 1 e sua lista de valores nas linhas seguintes.
    Saída: os bytes do arquivo .xlsx.
    """
    # Fase 1: workbook + aba.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = aba
    # Fase 2: para cada coluna, cabeçalho na linha 1 e valores a partir da linha 2.
    for c, (nome, valores) in enumerate(colunas.items(), start=1):
        ws.cell(row=1, column=c, value=nome)
        for r, valor in enumerate(valores, start=2):
            ws.cell(row=r, column=c, value=valor)
    # Saída: serializa em bytes.
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
