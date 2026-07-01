"""Parser da planilha Anexo V — aba `Preenchimento` (Bloco D1, §7).

Por que existe: a validação real precisa transformar o `.xlsx` enviado (bytes) em
linhas estruturadas antes de aplicar as regras. Este módulo isola essa leitura: abre a
planilha, exige a aba `Preenchimento`, mapeia as colunas **por nome do cabeçalho** (robusto
a reordenação; cabeçalho na linha 2, dados a partir da linha 3), e devolve só as linhas
**preenchidas pelo usuário** (as que têm ODI e/ou UC), cada uma com o número real da linha
(para o `loc` "L{n}"). Também oferece leitura **defensiva** de data e coordenada, já que
não há planilha real de exemplo (spec risco #2).

Erros estruturais (não-.xlsx, aba ausente) viram `PlanilhaInvalida` (→ HTTP 400 na rota).
A guarda de "0 linhas de dados" fica na camada de validação (§7), pois é erro de conteúdo,
não de formato.
"""

# `io.BytesIO` embrulha os bytes recebidos como um arquivo para o openpyxl.
import io
# `datetime` interpreta datas (texto ou serial do Excel).
import datetime
# `Path` resolve o caminho do modelo (para a aba Dominios).
from pathlib import Path
# `openpyxl` lê o .xlsx.
import openpyxl

# Nome da aba obrigatória e posições de cabeçalho/dados (§7).
_ABA = "Preenchimento"
_ABA_DOMINIOS = "Dominios"
_LINHA_CABECALHO = 2
# Modelo oficial (aba Dominios). Fica em `manuais/` (fora do git; presente no servidor).
_MODELO_PADRAO = (
    Path(__file__).resolve().parent.parent
    / "manuais" / "Anexo V - Planilha - Painel de Monitoramento - MME-CC_UF.xlsx"
)
# Colunas-chave que definem uma "linha de dados" (§7: tem ODI e/ou UC).
_COL_ODI = "Número ODI"
_COL_UC = "Número da Unidade Consumidora"


class PlanilhaInvalida(Exception):
    """Erro estrutural da planilha (não-.xlsx ou sem a aba `Preenchimento`) → 400."""

    def __init__(self, mensagem):
        # Guarda a mensagem legível para a resposta HTTP.
        super().__init__(mensagem)
        self.mensagem = mensagem


def _vazio(valor):
    """Diz se uma célula está vazia (None ou string só com espaços).

    Entrada: `valor` (qualquer).
    Saída: True se vazia; False caso contrário.
    """
    # None ou string em branco contam como vazio.
    return valor is None or (isinstance(valor, str) and valor.strip() == "")


def ler_preenchimento(conteudo):
    """Lê a aba `Preenchimento` e devolve as linhas de dados mapeadas por cabeçalho.

    Entrada: `conteudo` (bytes do .xlsx).
    Fase 1: abre o workbook; falha → `PlanilhaInvalida` (não-.xlsx).
    Fase 2: exige a aba `Preenchimento`; ausente → `PlanilhaInvalida`.
    Fase 3: lê o cabeçalho (linha 2) como mapa índice→nome.
    Fase 4: percorre a partir da linha 3; monta cada registro por nome, mantém só as
            linhas com ODI e/ou UC, anexando `_linha` (nº real na planilha).
    Saída: lista de dicts (uma por linha de dados), possivelmente vazia.
    """
    # Fase 1: abre em modo somente-leitura (eficiente p/ até 50k linhas) e só valores.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        # Qualquer falha de abertura = arquivo não é um .xlsx válido.
        raise PlanilhaInvalida("Arquivo não é um .xlsx válido.")
    # Fase 2: a aba Preenchimento é obrigatória.
    if _ABA not in wb.sheetnames:
        wb.close()
        raise PlanilhaInvalida(f"Aba '{_ABA}' não encontrada na planilha.")
    ws = wb[_ABA]
    # Fase 3+4: itera as linhas (values_only), separando cabeçalho e dados.
    cabecalho = None
    linhas = []
    for numero_linha, valores in enumerate(ws.iter_rows(values_only=True), start=1):
        # Antes da linha do cabeçalho (linha 1 = cabeçalho de seção) → ignora.
        if numero_linha < _LINHA_CABECALHO:
            continue
        # Linha do cabeçalho real → monta o mapa índice(0-based)→nome.
        if numero_linha == _LINHA_CABECALHO:
            cabecalho = {
                i: str(v).strip()
                for i, v in enumerate(valores)
                if v is not None and str(v).strip()
            }
            continue
        # Linha de dados: monta o registro por nome de coluna.
        registro = {
            nome: (valores[i] if i < len(valores) else None)
            for i, nome in cabecalho.items()
        }
        # Só conta como dado se tiver ODI e/ou UC (§7).
        if _vazio(registro.get(_COL_ODI)) and _vazio(registro.get(_COL_UC)):
            continue
        # Anexa o número real da linha (para o `loc` "L{n}").
        registro["_linha"] = numero_linha
        linhas.append(registro)
    wb.close()
    # Planilha sem cabeçalho na linha 2 = estrutura inválida.
    if cabecalho is None:
        raise PlanilhaInvalida(f"Aba '{_ABA}' sem cabeçalho na linha {_LINHA_CABECALHO}.")
    # Saída: linhas de dados (pode ser vazia — a guarda de 'sem dados' é da validação).
    return linhas


def ler_dominios(conteudo):
    """Lê a aba `Dominios` do modelo → dict {coluna: [valores válidos]} (Bloco D2, §7).

    Por que existe: as regras de domínio (D3) precisam das listas válidas (Tipo de
    Atendimento, UF, Sim/Não, Tipo de Comunidade, Enquadramento) — que são a fonte única
    do próprio modelo, não hardcoded.

    Entrada: `conteudo` (bytes do .xlsx do modelo).
    Fase 1: abre; falha → `PlanilhaInvalida`.
    Fase 2: exige a aba `Dominios`.
    Fase 3: linha 1 = cabeçalhos; cada coluna acumula os valores não-vazios abaixo.
    Saída: dict {nome_coluna: [valores]}.
    """
    # Fase 1: abre o workbook.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        raise PlanilhaInvalida("Arquivo não é um .xlsx válido.")
    # Fase 2: aba Dominios obrigatória.
    if _ABA_DOMINIOS not in wb.sheetnames:
        wb.close()
        raise PlanilhaInvalida(f"Aba '{_ABA_DOMINIOS}' não encontrada no modelo.")
    ws = wb[_ABA_DOMINIOS]
    # Fase 3: cabeçalhos (linha 1) e acúmulo dos valores por coluna.
    cabecalho = None
    colunas = {}
    for numero_linha, valores in enumerate(ws.iter_rows(values_only=True), start=1):
        # Linha 1: nomes das colunas de domínio.
        if numero_linha == 1:
            cabecalho = {
                i: str(v).strip()
                for i, v in enumerate(valores)
                if v is not None and str(v).strip()
            }
            colunas = {nome: [] for nome in cabecalho.values()}
            continue
        # Demais linhas: acumula o valor não-vazio de cada coluna.
        for i, nome in cabecalho.items():
            v = valores[i] if i < len(valores) else None
            if v is not None and str(v).strip():
                colunas[nome].append(str(v).strip())
    wb.close()
    # Saída: listas de domínio por coluna.
    return colunas


# Cache dos domínios do modelo (lidos uma vez; o modelo muda raramente).
_dominios_singleton = None


def obter_dominios(caminho=None):
    """Devolve os domínios do modelo oficial (cache), lendo a aba `Dominios`.

    Entrada: `caminho` (Path/str do .xlsx; None usa o modelo padrão em `manuais/`).
    Fase 1: na 1ª chamada, lê o arquivo e parseia; depois reaproveita.
    Saída: dict {coluna: [valores]}.
    """
    # Reatribui o cache de módulo.
    global _dominios_singleton
    # Fase 1: carrega sob demanda.
    if _dominios_singleton is None:
        alvo = Path(caminho) if caminho is not None else _MODELO_PADRAO
        with open(alvo, "rb") as fh:
            _dominios_singleton = ler_dominios(fh.read())
    # Saída: domínios cacheados.
    return _dominios_singleton


def normalizar_data(valor):
    """Interpreta uma data de energização de forma defensiva (§7, risco #2).

    Entrada: `valor` — datetime/date (serial do Excel) ou texto `DD/MM/AAAA`/`AAAA-MM-DD`.
    Fase 1: já é datetime/date → devolve a data.
    Fase 2: texto → tenta os formatos aceitos.
    Saída: `datetime.date` ou None se não reconhecer.
    """
    # Vazio → sem data.
    if valor is None:
        return None
    # Fase 1: serial do Excel chega como datetime/date.
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    # Fase 2: texto nos formatos aceitos.
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    # Saída: não reconhecida.
    return None


def normalizar_coordenada(valor):
    """Interpreta latitude/longitude de forma defensiva (decimal com `,` ou `.`).

    Entrada: `valor` — número ou texto (`-3,3018` ou `-3.3018`).
    Fase 1: número → float direto.
    Fase 2: texto → troca `,` por `.` e converte.
    Saída: float ou None se não for numérico.
    """
    # Vazio → sem coordenada.
    if valor is None:
        return None
    # Fase 1: já é número.
    if isinstance(valor, (int, float)):
        return float(valor)
    # Fase 2: texto com vírgula ou ponto decimal.
    texto = str(valor).strip().replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        # Saída: não numérico.
        return None
